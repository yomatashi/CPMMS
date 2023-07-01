from PySide6.QtCore import QRegularExpression, Qt
from PySide6.QtGui import QRegularExpressionValidator
from DB.connectionDB import FirebaseAccessor, FirebaseMutator
from PySide6.QtWidgets import QFileDialog, QHeaderView, QTableWidgetItem
from PySide6.QtWidgets import QMessageBox
import os
import pandas as pd
import openpyxl

def addItemScreen(self):
    self.stackedWidget.setCurrentWidget(self.addItem)

    # clear screen
    self.input_barcode.clear()
    self.input_item_name.clear()
    self.input_item_category.clear()
    self.input_price.clear()
    self.input_stock.clear()

    self.input_barcode.setPlaceholderText("Barcode ID")
    regexDigit = QRegularExpression(r"\d*")
    validator = QRegularExpressionValidator(regexDigit)
    self.input_barcode.setValidator(validator)
    self.input_item_name.setPlaceholderText("Item name")
    self.input_item_category.setPlaceholderText("Item category")
    self.input_price.setPlaceholderText("Price (RM)")
    regexAmnt = QRegularExpression(r'^[0-9]+(\.[0-9]{1,2})?$')
    validator2 = QRegularExpressionValidator(regexAmnt)
    self.input_price.setValidator(validator2)
    self.input_stock.setPlaceholderText("Stock")
    self.input_stock.setValidator(validator)

def listItemScreen(self):
    self.stackedWidget.setCurrentWidget(self.listItem)
    self.searchItem.clear()
    tableInv = self.tbl_itemList
    tableInv.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    tableInv.horizontalHeader().setMinimumSectionSize(100)
    message = ""
    isLowStock = 0

    # display inventory list
    get_inventory = FirebaseAccessor('Inventory').read_all()
    tableInv.setRowCount(len(get_inventory))
    row = 0
    for item in get_inventory:
        barcode_item = QTableWidgetItem(item['barcode'])
        tableInv.setItem(row, 0, barcode_item)

        name_item = QTableWidgetItem(item['itemName'])
        tableInv.setItem(row, 1, name_item)

        category_name = QTableWidgetItem(item['category'])
        tableInv.setItem(row, 2, category_name)

        price_item = QTableWidgetItem(str("{:.2f}".format(item['price'])))
        price_item.setTextAlignment(Qt.AlignCenter)
        tableInv.setItem(row, 3, price_item)

        stock_item = QTableWidgetItem(str(item['stock']))
        stock_item.setTextAlignment(Qt.AlignCenter)
        tableInv.setItem(row, 4, stock_item)

        if(item['stock'] < 10):
            message += item['itemName'] + "\n"
            isLowStock = 1
        row += 1

    if(isLowStock == 1):
        QMessageBox.warning(self, 'Warning! Low stock alert.', message, QMessageBox.Ok)
    
def browseExcel(self):
    doc_dir = os.path.expanduser("~/Documents")
    fname = QFileDialog.getOpenFileName(self, "Open File", doc_dir, "Excel Files (*.xlsx *.xls)")

    if fname:
        self.input_excel_name.setText(fname[0])

def importExcelScreen(self):
    self.stackedWidget.setCurrentWidget(self.importItem)
    self.input_excel_name.clear()

def uploadExcel(self):
        excelDir = self.input_excel_name.text()
        if not excelDir:
            QMessageBox.warning(self, "No file attached", "Please upload an excel file of inventory data", QMessageBox.Ok)
        else:
            df = pd.read_excel(excelDir, dtype={'BARCODE ID':str})
            Inventory = FirebaseMutator('Inventory')
            for index, row in df.iterrows():
                barcodeID = row['BARCODE ID']
                name = row['DESCRIPTION (ITEM NAME)']
                category = row['ITEM CATEGORY']
                price = row['SALE PRICE (RM)']
                stock = row['IN STOCK']
                Inventorydata = {'barcode': barcodeID, 'itemName': name, 'category': category ,'price': price, 'stock': stock}
                Inventory.create(Inventorydata, barcodeID)
                
            ret = QMessageBox.information(self, 'Success', 'The inventory data from the selected excel file has been added into database.', QMessageBox.Ok)
            if ret == QMessageBox.Ok:
                self.input_excel_name.clear()

def downloadCV(self):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventory Data"
    tableInv = self.tbl_itemList

    column_headers = ["BARCODE ID", "DESCRIPTION (ITEM NAME)", "SALE PRICE (RM)", "IN STOCK", "ITEM CATEGORY"]
    for col_index, header in enumerate(column_headers, start=1):
        sheet.cell(row=1, column=col_index, value=header)
    
    column_order = [0,1,3,4,2]
    for row in range(tableInv.rowCount()):
        for col, col_index in enumerate(column_order, start=1):
            item_text = tableInv.item(row, col_index).text()
            sheet.cell(row=row + 2, column=col, value=item_text)

    doc_dir = os.path.expanduser("~/Documents").replace("\\", "/")
    QMessageBox.information(self, 'Successfully downloaded', 'Inventory data has successfully downloaded as Excel file.\nFile location: '+doc_dir, QMessageBox.Ok)
    file_path = os.path.join(doc_dir, "Consult_Pharmacy_Inventory_Data.xlsx")
    workbook.save(file_path)

def createItem(self):
    barcodeID = self.input_barcode.text()
    name = self.input_item_name.text()
    category = self.input_item_category.text()
    price = self.input_price.text()
    stock = self.input_stock.text()

    if not all([barcodeID, name, category, price, stock]):
        QMessageBox.warning(self, "Error", "Please fill in the empty field(s)", QMessageBox.Ok)
    else:
        Inventory = FirebaseMutator('Inventory')
        Inventorydata = {'barcode': barcodeID, 'itemName': name, 'category': category, 'price': float(price), 'stock': int(stock)}
        Inventory.create(Inventorydata, barcodeID)

        ret = QMessageBox.information(self, 'Success', 'Item info has been added into database.', QMessageBox.Ok)
        if ret == QMessageBox.Ok:
            self.input_barcode.clear()
            self.input_item_name.clear()
            self.input_item_category.clear()
            self.input_price.clear()
            self.input_stock.clear()

def searchItem(self):
    tableInv = self.tbl_itemList
    tableInv.clearContents()
    tableInv.setRowCount(0)

    # display search item list
    search_input = self.searchItem.text()
    get_item = FirebaseAccessor('Inventory').read_all()

    lenCounter = 0
    row = 0
    for item in get_item:
        if search_input in item['barcode'] or search_input.lower() in item['itemName'].lower():
            lenCounter += 1
            tableInv.setRowCount(lenCounter)
            barcode_item = QTableWidgetItem(item['barcode'])
            tableInv.setItem(row, 0, barcode_item)

            name_item = QTableWidgetItem(item['itemName'])
            tableInv.setItem(row, 1, name_item)

            category_name = QTableWidgetItem(item['category'])
            tableInv.setItem(row, 2, category_name)

            price_item = QTableWidgetItem(str("{:.2f}".format(item['price'])))
            price_item.setTextAlignment(Qt.AlignCenter)
            tableInv.setItem(row, 3, price_item)

            stock_item = QTableWidgetItem(str(item['stock']))
            stock_item.setTextAlignment(Qt.AlignCenter)
            tableInv.setItem(row, 4, stock_item)
            row += 1

def show_item_details(self):
    selected_items = self.tbl_itemList.selectedItems()
    if len(selected_items) == 0:
        return
    row = selected_items[0].row()
    selected_data = self.tbl_itemList.item(row, 0).text() if self.tbl_itemList.item(row, 0) is not None else ""
    self.stackedWidget.setCurrentWidget(self.editItem)
    itemInfo = FirebaseAccessor('Inventory').read(selected_data)

    self.input_barcodeID_edit.setText(itemInfo['barcode'])
    regexDigit = QRegularExpression(r"\d*")
    validator = QRegularExpressionValidator(regexDigit)
    self.input_barcodeID_edit.setValidator(validator)
    self.input_itemName_edit.setText(itemInfo['itemName'])
    self.input_itemCategory_edit.setText(itemInfo['category'])
    self.input_price_edit.setText(str("{:.2f}".format(itemInfo['price'])))
    regexAmnt = QRegularExpression(r'^[0-9]+(\.[0-9]{1,2})?$')
    validator2 = QRegularExpressionValidator(regexAmnt)
    self.input_price_edit.setValidator(validator2)
    self.input_stock_edit.setText(str(itemInfo['stock']))
    self.input_stock_edit.setValidator(validator)

def delete_item(self):
    item_id = self.input_barcodeID_edit.text()
    ret = QMessageBox.warning(self, "Delete confirmation", "Are you sure you want to delete this item?", QMessageBox.Yes | QMessageBox.No)
    if ret == QMessageBox.Yes:
        FirebaseMutator("Inventory").delete(item_id)
        self.stackedWidget.setCurrentWidget(self.listItem)
        QMessageBox.information(self, "Success", "Selected item data deleted!\nPlease refresh inventory list", QMessageBox.Ok)

def update_item(self):
    barcode = self.input_barcodeID_edit.text()
    itemName = self.input_itemName_edit.text()
    category = self.input_itemCategory_edit.text()
    price = float(self.input_price_edit.text())
    stock = int(self.input_stock_edit.text())

    update_item = FirebaseMutator('Inventory')
    update_data = {'barcode': barcode, 'itemName': itemName, 'category': category, 'price': price, 'stock': stock}
    update_item.update(barcode, update_data)
    QMessageBox.information(self, "Updated", "Item information successfully updated.", QMessageBox.Ok)